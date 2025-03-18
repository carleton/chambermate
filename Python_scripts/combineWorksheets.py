import xlrd
import xlwt
import os
import openpyxl
from openpyxl.utils import get_column_letter

import sys
import glob
import csv
from xlsxwriter.workbook import Workbook
import subprocess
from os.path import expanduser


# Make sure to pass the name of the folder you want to combine as the first argument.
# This folder should only have the files containing the output data
# At the moment of writing this, the folder is named 'output'
try:
	folderName = sys.argv[1]
except:
	print('Please enter the name of the folder as the first argument')

#Get rid of the slash in the folder name
folderName = folderName[:len(folderName) - 1]

#Copy the template file
subprocess.call(["cp", "temp.xlsx", folderName + ".xlsx"])

#Create a new workbook book to write to
wb2 = openpyxl.load_workbook(filename = folderName + ".xlsx")

#Let a be all the files in the folder
a = os.listdir(folderName)

#Go through all the files and remove the ones that start with a ~ or .
#These are the old versions of files or hidden files
b = []
for x in a:
	if(x[:1] == "~" or x[:1] == "." or x == "Icon"):
		b.append(x)
for x in b:
	a.remove(x)

for csvfile in glob.glob(os.path.join('.', '*.csv')):
	workbook = Workbook(csvfile[:-4] + '.xlsx')
	worksheet = workbook.add_worksheet()
	with open(csvfile, 'rt', encoding='utf8') as f:
		reader = csv.reader(f)
		for r, row in enumerate(reader):
			for c, col in enumerate(row):
				worksheet.write(r, c, col)
	workbook.close()

#Sort by the female number, found by taking the first characters of the file name


key_list = lambda x: int(x[:2]) if str(x[:2]).isdigit() else int(x[:1])
a = sorted(a, key=key_list)



#For each file copy its contents into a page of the workbook
workbookNumber = 0
for filename in a: 
	workbookNumber = workbookNumber + 1
	print(filename)
		
	#Double check that we aren't copying in old files or hidden files
	if(filename == '.DS_Store' or filename[:1] == "~"):
		continue
	elif(filename[-4:] == ".csv"):
		csv_filename = filename
		f = open(folderName + "/" + csv_filename)
		reader = csv.reader(f)
		ws = wb2.create_sheet(str(workbookNumber))
		for row_index, row in enumerate(reader):
			for column_index, cell in enumerate(row):
				column_letter = get_column_letter((column_index + 1))
				s = cell
				try:
					s=float(s)
				except ValueError:
					pass

				ws.cell(row_index + 1, column_index + 1).value = s
				if(ws.cell(row_index + 1, column_index + 1).value == "&nbsp;"):
					ws[ws.cell(row_index + 1, column_index + 1).coordinate].value = ""
		continue

	#Create the new page in the workbook with name based on the female number in the file
	else:
		wb1 = openpyxl.load_workbook(filename=folderName + "/" +filename)

	ws1 = wb1.worksheets[0]
	ws2 = wb2.create_sheet(str(workbookNumber))

	#Copy every cell except ones containing "&nbsp;"
	for row in ws1:
		for cell in row:
			ws2[cell.coordinate].value = cell.value
			if(cell.value == "&nbsp;"):
				ws2[cell.coordinate].value = ""
wb2.remove_sheet(wb2.get_sheet_by_name('Sheet'))
#Save the workbook as an excel file
wb2.save(filename=folderName + ".xlsx")
